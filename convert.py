# -*- coding: utf-8 -*-
"""
Converte a planilha 'ACOMPANHAMENTO PROFISCO II' (aba 'Visão Geral') em data.json,
no schema consumido pelo app (index.html).

Uso:
    python convert.py "Copia de ACOMPANHAMENTO PROFISCO II - vs 25_06_2026.xlsx"
Se nenhum caminho for passado, procura o primeiro .xlsx na pasta atual.
"""
import sys, glob, json, datetime, os

def find_file():
    if len(sys.argv) > 1:
        return sys.argv[1]
    xs = glob.glob("*.xlsx")
    if not xs:
        print("Nenhum .xlsx encontrado."); sys.exit(1)
    return xs[0]

# Mapeamento fixo de colunas (cabeçalho em 2 linhas na planilha)
COLS = [
    ("id",              "ID",                         "id"),
    ("produto",         "Produto",                    "title"),
    ("areas",           "Áreas",                      "text"),
    ("situacao",        "Situação",                   "status"),
    ("progresso",       "Progresso (%)",              "percent0"),
    ("orc_bid",         "Orçamento BID (US$)",        "money"),
    ("orc_contra",      "Contrapartida (US$)",        "money"),
    ("orc_total",       "Total (US$)",                "money"),
    ("participacao",    "Participação (%)",           "percent"),
    ("orc_rs",          "Orçamento R$ (Estim. SEFIN)","money"),
    ("processo_sei",    "Processo SEI",               "longtext"),
    ("contrato",        "Contrato",                   "text"),
    ("proximos",        "Próximos Passos / Entregáveis","longtext"),
    ("lider",           "Líder",                      "longtext"),
    ("gerente",         "Gerente",                    "longtext"),
    ("inicio",          "Início",                     "date"),
    ("conclusao",       "Conclusão",                  "date"),
    ("tap",             "TAP",                        "text"),
    ("eap",             "EAP",                        "text"),
    ("sgdi",            "SGDI",                       "select"),
    ("sgdi_obs",        "SGDI - Observações",         "longtext"),
]
STATUS_OPTIONS = ["Não iniciado", "Em Andamento", "Concluído", "Atrasado", "Bloqueado", "NA"]
SGDI_OPTIONS = ["", "Sim", "Não"]

def norm(v):
    if v is None: return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v

def default_progress(situacao):
    s = str(situacao).strip().lower()
    if s.startswith("conclu"): return 100
    if "andamento" in s: return 40
    if "atras" in s: return 40
    if "bloque" in s: return 20
    return 0

def main():
    import openpyxl
    f = find_file()
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(3, ws.max_row + 1):
        c = lambda i: norm(ws.cell(r, i).value)
        idv, comp, prod = c(2), c(1), c(3)
        # linha válida: tem ID ou é a linha do componente (linha 3)
        if not idv and not comp:
            continue
        if not idv and comp:
            idv = str(comp).strip().split(".")[0]  # "3"
            prod = comp
        situacao = c(10) or "Não iniciado"
        sgdi = "Sim" if str(c(20)).strip() else ("Não" if str(c(21)).strip() else "")
        rec = {
            "id": str(idv),
            "produto": prod,
            "areas": c(4),
            "situacao": situacao,
            "progresso": default_progress(situacao),
            "orc_bid": c(5),
            "orc_contra": c(6),
            "orc_total": c(7),
            "participacao": c(8),
            "orc_rs": c(9),
            "processo_sei": c(11),
            "contrato": c(12),
            "proximos": c(13),
            "lider": c(14),
            "gerente": c(15),
            "inicio": c(16),
            "conclusao": c(17),
            "tap": c(18),
            "eap": c(19),
            "sgdi": sgdi,
            "sgdi_obs": c(22),
            "level": str(idv).count("."),
            "_history": [],
        }
        rows.append(rec)

    data = {
        "meta": {
            "title": "Acompanhamento PROFISCO II — Componente 3",
            "subtitle": "Administração Financeira e Gasto Público — SEEC/DF",
            "source": os.path.basename(f),
            "updated": datetime.date.today().isoformat(),
        },
        "columns": [{"key": k, "label": l, "type": t} for (k, l, t) in COLS],
        "statusOptions": STATUS_OPTIONS,
        "sgdiOptions": SGDI_OPTIONS,
        "rows": rows,
    }
    with open("data.json", "w", encoding="utf-8") as out:
        json.dump(data, out, ensure_ascii=False, indent=2)
    print("data.json gerado com %d linhas." % len(rows))

if __name__ == "__main__":
    main()
