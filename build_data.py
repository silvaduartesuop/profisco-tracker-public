# -*- coding: utf-8 -*-
"""
Script ÚNICO: converte a planilha 'ACOMPANHAMENTO PROFISCO II' (aba 'Visão Geral')
em data.json e, no mesmo passo, gera a versão anonimizada (sem nomes/matrículas).

Saídas:
  - data.json         -> base COMPLETA (uso interno / repositório privado)
  - data.public.json  -> base ANONIMIZADA (site público; governança omitida)

Uso:
  python build_data.py "ACOMPANHAMENTO PROFISCO II - vs XX_XX_XXXX.xlsx"
  (sem argumento, usa o primeiro .xlsx da pasta)

Requer: pip install openpyxl
"""
import sys, glob, json, re, datetime, os

COLS = [
    ("id","ID","id"),
    ("produto","Produto","title"),
    ("areas","Áreas","text"),
    ("situacao","Situação","status"),
    ("progresso","Progresso (%)","percent0"),
    ("orc_bid","Orçamento BID (US$)","money"),
    ("orc_contra","Contrapartida (US$)","money"),
    ("orc_total","Total (US$)","money"),
    ("participacao","Participação (%)","percent"),
    ("orc_rs","Orçamento R$ (Estim. SEFIN)","money"),
    ("processo_sei","Processo SEI","longtext"),
    ("contrato","Contrato","text"),
    ("proximos","Próximos Passos / Entregáveis","longtext"),
    ("lider","Líder","longtext"),
    ("gerente","Gerente","longtext"),
    ("inicio","Início","date"),
    ("conclusao","Conclusão","date"),
    ("tap","TAP","text"),
    ("eap","EAP","text"),
    ("sgdi","SGDI","select"),
    ("sgdi_obs","SGDI - Observações","longtext"),
]
STATUS_OPTIONS = ["Não iniciado","Em Andamento","Concluído","Atrasado","Bloqueado","NA"]
SGDI_OPTIONS = ["","Sim","Não"]
PII_FIELDS = {"lider","gerente"}
MAT = re.compile(r'matr[íi]cula[^;\n]*', re.I)
NUM = re.compile(r'\bn[ºo°]?\s*\d{1,3}[.\d]*\-?[\dxX]\b')

def norm(v):
    if v is None: return ""
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

def default_progress(s):
    s=str(s).strip().lower()
    if s.startswith("conclu"): return 100
    if "andamento" in s or "atras" in s: return 40
    if "bloque" in s: return 20
    return 0

def read_rows(path):
    import openpyxl
    wb=openpyxl.load_workbook(path, data_only=True); ws=wb[wb.sheetnames[0]]
    rows=[]
    for r in range(3, ws.max_row+1):
        c=lambda i: norm(ws.cell(r,i).value)
        idv, comp, prod = c(2), c(1), c(3)
        if not idv and not comp: continue
        if not idv and comp:
            idv=str(comp).strip().split(".")[0]; prod=comp
        situ=c(10) or "Não iniciado"
        sgdi="Sim" if str(c(20)).strip() else ("Não" if str(c(21)).strip() else "")
        rows.append({
            "id":str(idv),"produto":prod,"areas":c(4),"situacao":situ,
            "progresso":default_progress(situ),
            "orc_bid":c(5),"orc_contra":c(6),"orc_total":c(7),"participacao":c(8),"orc_rs":c(9),
            "processo_sei":c(11),"contrato":c(12),"proximos":c(13),
            "lider":c(14),"gerente":c(15),"inicio":c(16),"conclusao":c(17),
            "tap":c(18),"eap":c(19),"sgdi":sgdi,"sgdi_obs":c(22),
            "level":str(idv).count("."),"ativo":True,"_history":[],
        })
    return rows

def wrap(rows, title, subtitle, source):
    return {
        "meta":{"title":title,"subtitle":subtitle,"source":source,
                "updated":datetime.date.today().isoformat()},
        "columns":[{"key":k,"label":l,"type":t} for (k,l,t) in COLS],
        "statusOptions":STATUS_OPTIONS,"sgdiOptions":SGDI_OPTIONS,"rows":rows,
    }

def anonymize(rows):
    import copy
    out=copy.deepcopy(rows)
    for r in out:
        for k,v in list(r.items()):
            if k in PII_FIELDS and isinstance(v,str) and v.strip():
                r[k]=""
            elif isinstance(v,str) and v:
                nv=NUM.sub("",MAT.sub("",v)).strip()
                if nv!=v: r[k]=nv
    return out

def main():
    src=sys.argv[1] if len(sys.argv)>1 else (glob.glob("*.xlsx") or [None])[0]
    if not src: print("Nenhum .xlsx encontrado."); sys.exit(1)
    name=os.path.basename(src)
    rows=read_rows(src)

    full=wrap(rows,"Acompanhamento PROFISCO II — Componente 3",
              "Administração Financeira e Gasto Público — SEEC/DF", name)
    with open("data.json","w",encoding="utf-8") as f: json.dump(full,f,ensure_ascii=False,indent=2)

    pub=wrap(anonymize(rows),"Acompanhamento PROFISCO II — Componente 3 (versão pública)",
             "Administração Financeira e Gasto Público — SEEC/DF · dados de governança (responsáveis) omitidos", name)
    with open("data.public.json","w",encoding="utf-8") as f: json.dump(pub,f,ensure_ascii=False,indent=2)

    # checagem de PII na versão pública
    blob=json.dumps(pub,ensure_ascii=False)
    assert not re.search(r'matr[íi]cula',blob,re.I), "PII remanescente!"
    print("OK: %d linhas -> data.json (completo) e data.public.json (anonimizado)." % len(rows))
    print("   * suba data.json no repo PRIVADO e data.public.json (renomeado para data.json) no repo PÚBLICO.")

if __name__=="__main__":
    main()
